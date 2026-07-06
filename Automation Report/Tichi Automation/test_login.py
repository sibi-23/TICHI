from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook


def test_login():

    workbook = load_workbook(r"C:\Users\sibic\Downloads\login-data.xlsx")  
    sheet = workbook.active

    driver = webdriver.Chrome()
    driver.maximize_window()

    wait = WebDriverWait(driver, 20)

    rows = sheet.max_row

    for i in range(2, rows + 1):
        email = sheet.cell(row=i, column=1).value
        password = sheet.cell(row=i, column=2).value
        if email is None or password is None:
            print(f"Skipping Row {i}")
            continue
        driver.get("https://tichi-app-webapp-stage.web.app/login")
        email_box = wait.until(
            EC.visibility_of_element_located((By.ID, "email"))
        )
        email_box.clear()
        email_box.send_keys(str(email))
        wait.until(
            EC.element_to_be_clickable(
                (By.XPATH, "/html/body/div[1]/div[2]/div/form/button")
            )
        ).click()
        password_box = wait.until(
            EC.visibility_of_element_located((By.ID, "password"))
        )
        password_box.clear()
        password_box.send_keys(str(password))
        wait.until(
            EC.element_to_be_clickable(
                (By.XPATH, "/html/body/div[1]/div[2]/div/form/button")
            )
        ).click()

    
        if "home" in driver.current_url:
            print(f"Row {i}: PASS")
            sheet.cell(row=i, column=3).value = "PASS"
        else:
            print(f"Row {i}: FAIL")
            sheet.cell(row=i, column=3).value = "FAIL"

    workbook.save("login-data.xlsx")
    driver.quit()
